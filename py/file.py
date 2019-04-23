import os
import shutil
import time
import openpyxl

path = "C:\\Users\\h.c.wu\\Desktop\\实时监控\\memory\\"
filename = "生产监控" + time.strftime("%Y%m%d", time.localtime()) + ".xlsx"

os.chdir(path)
shutil.copy(path + "生产监控Demo.xlsx",
            path + filename)

from openpyxl import Workbook
wb = openpyxl.load_workbook(filename)
print(wb.sheetnames)

ws = wb["内网"]
for row in ws.iter_rows(values_only=True):
    print(row)
